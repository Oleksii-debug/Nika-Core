from __future__ import annotations

import csv
import hashlib
from importlib import metadata
from io import BytesIO, StringIO

import pytest
from docx import Document
from openpyxl import load_workbook

from nika_core.packaging.notices import RUNTIME_DISTRIBUTIONS
from nika_core.research.models import FreshnessState, ResearchEvidence, SourceKind
from nika_core.research.report_exports import ResearchReportExporter, ResearchReportFormat
from nika_core.research.review import (
    AccessibleResearchReport,
    ResearchCard,
    ResearchReview,
    ResearchReviewState,
)


def _report() -> AccessibleResearchReport:
    card = ResearchCard(
        ordinal=0,
        document_id="doc-1",
        title="=1+1 Українська <можливість>",
        snippet="Грант & навчання <script>alert(1)</script>",
        rank=-1.25,
        why_matched="Literal match: грант",
        evidence=(
            ResearchEvidence(
                source_id="source-1",
                source_kind=SourceKind.HTTP,
                locator="https://example.org/?a=1&b=2",
                observed_at="2026-08-20T07:00:00+00:00",
                freshness=FreshnessState.CURRENT,
            ),
        ),
        review=ResearchReview(
            workspace_id="ws",
            document_id="doc-1",
            state=ResearchReviewState.SAVED,
            note="  @SUM(A1:A2) перевірити вручну",
            updated_at="2026-08-20T07:10:00+00:00",
        ),
    )
    return AccessibleResearchReport(
        result_set_id="../results/unsafe id",
        workspace_id="ws",
        query="грант & навчання",
        created_at="2026-08-20T07:00:00+00:00",
        cards=(card,),
        text=(
            "Research results\n"
            "Query: грант & навчання\n"
            "Results: 1\n"
            "Result 1: =1+1 Українська <можливість>\n"
        ),
    )


def test_txt_export_is_utf8_and_returns_safe_leaf_filename() -> None:
    rendered = ResearchReportExporter().render(_report(), ResearchReportFormat.TXT)

    assert rendered.filename == "research-results-results-unsafe-id.txt"
    assert "/" not in rendered.filename
    assert "\\" not in rendered.filename
    assert rendered.media_type == "text/plain; charset=utf-8"
    assert rendered.content.decode() == _report().text
    assert rendered.sha256 == hashlib.sha256(rendered.content).hexdigest()


def test_csv_preserves_review_provenance_and_blocks_formula_injection() -> None:
    rendered = ResearchReportExporter().render(_report(), ResearchReportFormat.CSV)
    rows = list(csv.DictReader(StringIO(rendered.content.decode())))

    assert len(rows) == 1
    row = rows[0]
    assert row["title"] == "'=1+1 Українська <можливість>"
    assert row["review_state"] == "saved"
    assert row["review_note"] == "'  @SUM(A1:A2) перевірити вручну"
    assert row["review_updated_at"] == "2026-08-20T07:10:00+00:00"
    assert row["source_id"] == "source-1"
    assert row["source_kind"] == "http"
    assert row["freshness"] == "current"
    assert row["locator"] == "https://example.org/?a=1&b=2"
    assert row["observed_at"] == "2026-08-20T07:00:00+00:00"


def test_html_is_semantic_and_escapes_untrusted_text() -> None:
    rendered = ResearchReportExporter().render(_report(), ResearchReportFormat.HTML)
    text = rendered.content.decode()

    assert '<html lang="und">' in text
    assert "<main>" in text
    assert "<h1>Research results</h1>" in text
    assert '<article aria-labelledby="result-1">' in text
    assert (
        '<h2 id="result-1">Result 1: =1+1 Українська &lt;можливість&gt;</h2>'
        in text
    )
    assert "Грант &amp; навчання &lt;script&gt;alert(1)&lt;/script&gt;" in text
    assert "<dt>Review updated</dt><dd>2026-08-20T07:10:00+00:00</dd>" in text
    assert "https://example.org/?a=1&amp;b=2" in text
    assert "<script>alert(1)</script>" not in text


def test_docx_has_heading_hierarchy_labels_review_and_provenance() -> None:
    rendered = ResearchReportExporter().render(_report(), ResearchReportFormat.DOCX)
    document = Document(BytesIO(rendered.content))
    paragraphs = [(paragraph.style.name, paragraph.text) for paragraph in document.paragraphs]

    assert ("Heading 1", "Research results") in paragraphs
    assert ("Heading 2", "Result 1: =1+1 Українська <можливість>") in paragraphs
    assert ("Heading 3", "Evidence") in paragraphs
    assert ("Normal", "Review: saved") in paragraphs
    assert ("Normal", "Review updated: 2026-08-20T07:10:00+00:00") in paragraphs
    assert ("Normal", "Source ID: source-1") in paragraphs
    assert ("Normal", "Location: https://example.org/?a=1&b=2") in paragraphs
    assert document.core_properties.title == "Research results"
    assert document.core_properties.author == "Nika Core"
    assert document.core_properties.created.isoformat() == "2026-08-20T07:00:00+00:00"
    assert document.core_properties.modified.isoformat() == "2026-08-20T07:00:00+00:00"


def test_xlsx_is_flat_accessible_and_blocks_formula_injection() -> None:
    rendered = ResearchReportExporter().render(_report(), ResearchReportFormat.XLSX)
    workbook = load_workbook(BytesIO(rendered.content), data_only=False)

    assert workbook.sheetnames == ["Metadata", "Results"]
    assert workbook["Metadata"]["A1"].value == "Field"
    assert workbook["Metadata"].freeze_panes == "A2"
    assert workbook["Results"].freeze_panes == "A2"
    assert not workbook["Metadata"].merged_cells.ranges
    assert not workbook["Results"].merged_cells.ranges
    headers = [cell.value for cell in workbook["Results"][1]]
    row = [cell.value for cell in workbook["Results"][2]]
    values = dict(zip(headers, row, strict=True))
    assert values["title"] == "'=1+1 Українська <можливість>"
    assert values["review_note"] == "'  @SUM(A1:A2) перевірити вручну"
    assert values["review_updated_at"] == "2026-08-20T07:10:00+00:00"
    assert values["source_id"] == "source-1"
    assert values["locator"] == "https://example.org/?a=1&b=2"
    assert workbook["Results"]["G2"].data_type == "s"
    assert workbook.properties.title == "Research results"
    assert workbook.properties.creator == "Nika Core"
    assert workbook.properties.created.isoformat() == "2026-08-20T07:00:00"
    assert workbook.properties.modified.isoformat() == "2026-08-20T07:00:00"


def test_all_formats_are_byte_deterministic_for_same_report() -> None:
    exporter = ResearchReportExporter()
    report = _report()

    for report_format in ResearchReportFormat:
        first = exporter.render(report, report_format)
        second = exporter.render(report, report_format)
        assert first.content == second.content
        assert first.sha256 == second.sha256


def test_office_export_rejects_non_iso_created_at() -> None:
    report = _report()
    invalid = AccessibleResearchReport(
        result_set_id=report.result_set_id,
        workspace_id=report.workspace_id,
        query=report.query,
        created_at="not-a-timestamp",
        cards=report.cards,
        text=report.text,
    )

    for report_format in (ResearchReportFormat.DOCX, ResearchReportFormat.XLSX):
        with pytest.raises(ValueError, match="ISO-8601"):
            ResearchReportExporter().render(invalid, report_format)


def test_exporter_rejects_untyped_format() -> None:
    with pytest.raises(TypeError, match="ResearchReportFormat"):
        ResearchReportExporter().render(_report(), "txt")  # type: ignore[arg-type]


@pytest.mark.parametrize(
    "distribution_name",
    ("python-docx", "openpyxl", "lxml", "et-xmlfile"),
)
def test_report_export_runtime_dependencies_have_notice_and_license_evidence(
    distribution_name: str,
) -> None:
    assert distribution_name in RUNTIME_DISTRIBUTIONS
    distribution = metadata.distribution(distribution_name)
    license_evidence = (
        distribution.metadata.get("License-Expression")
        or distribution.metadata.get("License")
        or "".join(
            value
            for value in distribution.metadata.get_all("Classifier", [])
            if value.startswith("License ::")
        )
    )
    assert distribution.version
    assert license_evidence.strip()
